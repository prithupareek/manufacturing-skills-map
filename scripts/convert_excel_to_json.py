#!/usr/bin/env python3
"""
Convert two Excel files into JSON data for the manufacturing skills heat map.

Task 1: SAE numeracy data -> data/numeracy.json
Task 2: AI data center projects -> data/demand.json
"""

import json
import os
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_PATH = "/Users/komalpareek/Documents/prithu_code/manufacturing_heatmap/manufacturing-skills-map"
DATA_DIR = os.path.join(BASE_PATH, "data")

SAE_FILE = "/Users/komalpareek/Downloads/SAE_website_dataset.c15d59a2d7e219fcb6d1.xlsx"
PROJECTS_FILE = "/Users/komalpareek/Downloads/ai_data_center_projects.xlsx"

STATE_ABBREV = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR', 'California': 'CA',
    'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE', 'District of Columbia': 'DC',
    'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI', 'Idaho': 'ID', 'Illinois': 'IL',
    'Indiana': 'IN', 'Iowa': 'IA', 'Kansas': 'KS', 'Kentucky': 'KY', 'Louisiana': 'LA',
    'Maine': 'ME', 'Maryland': 'MD', 'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN',
    'Mississippi': 'MS', 'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE', 'Nevada': 'NV',
    'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM', 'New York': 'NY',
    'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH', 'Oklahoma': 'OK', 'Oregon': 'OR',
    'Pennsylvania': 'PA', 'Rhode Island': 'RI', 'South Carolina': 'SC', 'South Dakota': 'SD',
    'Tennessee': 'TN', 'Texas': 'TX', 'Utah': 'UT', 'Vermont': 'VT', 'Virginia': 'VA',
    'Washington': 'WA', 'West Virginia': 'WV', 'Wisconsin': 'WI', 'Wyoming': 'WY',
    'Puerto Rico': 'PR', 'Guam': 'GU', 'Virgin Islands': 'VI',
    'American Samoa': 'AS', 'Northern Mariana Islands': 'MP',
}

COORDS = {
    'Northern Virginia': (39.0437, -77.4875),
    'Childress': (34.4265, -100.2040),
    'St. Joseph': (41.6764, -86.2520),
    'Fulton': (33.7902, -84.3885),
    'Dorchester': (33.0765, -80.4074),
    'Chesterfield': (37.3776, -77.5058),
    'Wilmington': (39.7391, -75.5398),
    'Licking': (40.0914, -82.4832),
}

# Suffixes to strip from county names, ordered longest-first so we match
# the most specific suffix before a shorter one.
COUNTY_SUFFIXES = [
    " City and Borough",
    " Census Area",
    " Municipality",
    " Borough",
    " Parish",
    " County",
    " city",          # Virginia independent cities end with " City"
    " City",          # catch title-case variant too
]


def clean_county_name(raw: str) -> str:
    """Strip geographic-type suffixes from a county name."""
    name = raw.strip()
    for suffix in COUNTY_SUFFIXES:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break
    return name.strip()


# ---------------------------------------------------------------------------
# Task 1 — numeracy.json
# ---------------------------------------------------------------------------

def build_numeracy():
    print("Reading SAE dataset ...")
    wb = openpyxl.load_workbook(SAE_FILE, read_only=True, data_only=True)
    ws = wb["County"]

    numeracy = {}
    skipped_grp = 0
    skipped_state = 0
    null_rate = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fips_raw = row[0]
        state_name = row[1]
        county_raw = row[2]
        grp = row[3]
        num_p1 = row[32]

        # Only keep the "all" group
        if grp != "all":
            skipped_grp += 1
            continue

        # FIPS: integer -> 5-digit zero-padded string
        fips = str(int(fips_raw)).zfill(5)

        # State abbreviation
        abbrev = STATE_ABBREV.get(state_name)
        if abbrev is None:
            skipped_state += 1
            continue

        # County name cleanup
        county = clean_county_name(str(county_raw))

        # Rate calculation
        rate = None
        if num_p1 is not None and num_p1 != "NA":
            try:
                val = float(num_p1)
                rate = round((1 - val) * 100, 1)
            except (ValueError, TypeError):
                pass

        if rate is None:
            null_rate += 1

        numeracy[fips] = {"rate": rate, "county": county, "state": abbrev}

    wb.close()

    # Write JSON
    out_path = os.path.join(DATA_DIR, "numeracy.json")
    with open(out_path, "w") as f:
        json.dump(numeracy, f, separators=(",", ":"))

    print(f"  Wrote {out_path}")
    print(f"  Total entries: {len(numeracy)}")
    print(f"  Skipped (non-all grpName): {skipped_grp}")
    print(f"  Skipped (unknown state): {skipped_state}")
    print(f"  Entries with null rate: {null_rate}")

    return numeracy


# ---------------------------------------------------------------------------
# Task 2 — demand.json
# ---------------------------------------------------------------------------

def build_demand():
    print("Reading AI data center projects ...")
    wb = openpyxl.load_workbook(PROJECTS_FILE, read_only=True, data_only=True)
    ws = wb["Projects"]

    demand = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        location_text = str(row[0]).strip()
        project_value_text = str(row[1]).strip()
        value_b = row[3]  # numeric, in billions

        # Match coordinates
        lat, lng = None, None
        matched_key = None
        for key, (lt, ln) in COORDS.items():
            if key in location_text:
                lat, lng = lt, ln
                matched_key = key
                break

        if lat is None:
            print(f"  WARNING: no coordinate match for '{location_text}'")
            continue

        # Demand value: billions * 1000
        demand_val = round(float(value_b) * 1000)

        # Short name: use the matched key as the basis, then append state
        # Extract state from the location text (last word or known pattern)
        name = matched_key

        demand.append({
            "name": location_text,
            "lat": lat,
            "lng": lng,
            "demand": demand_val,
            "description": project_value_text,
        })

    wb.close()

    out_path = os.path.join(DATA_DIR, "demand.json")
    with open(out_path, "w") as f:
        json.dump(demand, f, indent=2)

    print(f"  Wrote {out_path}")
    print(f"  Total entries: {len(demand)}")

    return demand


# ---------------------------------------------------------------------------
# Task 3 — Validation
# ---------------------------------------------------------------------------

def validate_numeracy(data):
    print("\n--- Numeracy Validation ---")
    n = len(data)
    print(f"  Entry count: {n}  (expect ~3142)")

    # Check all keys are 5-digit FIPS
    bad_fips = [k for k in data if len(k) != 5 or not k.isdigit()]
    print(f"  Bad FIPS keys: {len(bad_fips)}")
    if bad_fips:
        print(f"    Examples: {bad_fips[:5]}")

    # Rate statistics
    rates = [v["rate"] for v in data.values() if v["rate"] is not None]
    nulls = n - len(rates)
    print(f"  Rates present: {len(rates)}, null: {nulls}")
    if rates:
        print(f"  Rate range: {min(rates):.1f} - {max(rates):.1f}")
        avg = sum(rates) / len(rates)
        print(f"  Rate mean:  {avg:.1f}")

    # Check plausibility (rates should be 0-100)
    out_of_range = [r for r in rates if r < 0 or r > 100]
    print(f"  Rates outside 0-100: {len(out_of_range)}")

    # Spot-check a few known FIPS
    for fips in ["01001", "06037", "48201", "36061"]:
        entry = data.get(fips)
        if entry:
            print(f"  Spot-check {fips}: {entry}")


def validate_demand(data):
    print("\n--- Demand Validation ---")
    print(f"  Entry count: {len(data)}  (expect 8)")
    for item in data:
        print(f"    {item['name'][:50]:50s}  lat={item['lat']:.4f}  lng={item['lng']:.4f}  demand={item['demand']}")

    # Check coordinates are in continental US range (roughly)
    for item in data:
        if not (24 < item["lat"] < 50 and -130 < item["lng"] < -60):
            print(f"  WARNING: coordinates out of CONUS range for {item['name']}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)

    numeracy = build_numeracy()
    demand = build_demand()

    validate_numeracy(numeracy)
    validate_demand(demand)

    print("\nDone.")
